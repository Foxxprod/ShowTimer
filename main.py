#Projet de prompteur pour le PCP N°... 2026


#Dev : Titouan Gallin
#Contact : titouan.gallin@gmail.com / +33 6 52 75 25 29

#--------------------------------------------------------------------------------------------------------------------

#Fonction a implementer :
# - Interface permetant de creer une emmision OKK
# - Interface de controle de l'emmision : choix de la durée, placement sur le timer de différents événements, lancement des timers... OKK
# - Pour chaques evenements TC, pouvoir envoyer un message OSC a chataigne OKK
# - Preview de l'affichage du prompteur dans la fenetre principale 
# - Saisie d'un message pour l'envoyer au prompteur, affichage en direct (en meme temps que l'on ecrit sur le clavier) OKK
# - Fenetre d'affichage du prompteur en double écran (avec fond pour chromakey ou lumakey) OKK
# - Pour chaques timer, on affiche au dessus un label avec un message du genre "Retour plateau dans..." OKK
# - Possibilitée d'affciher en meme temps plusieurs timers OKK -- RETIREE ENFAIT DANS LA DERNIERE VERSION
# - Pouvoir faire clignoter le texte du prompteur quand un nouveau message a été saisie OKK
# - Rendre plus visible le timers quand le temps arrive au bout (changement de couleur, clignotement...) OKK
# - Avoir contour sur les texte du prompteur -- VOIR COMMENT FAIRE CAR PAS POSSIBLE SUR QT OKK
# - Pouvoir importer depuis excel un conducteur OKK
# - pas pouvoir modifier tableau d'ajout des shows OKK



#--------------------------------------------------------------------------------------------------------------------
import database, utils, osc, logger
import sys, threading, os, openpyxl
from timeline import TimelineWidget
from ui.ui_main import Ui_MainWindow
from ui.ui_operator import Ui_Dialog as Ui_OperatorDialog
from prompter import PrompterWindow
from utils import PrompterSettingsDialog, export_show_to_pdf

from PySide6.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QPushButton, QLabel, QLineEdit, QTextEdit, QHBoxLayout, QGridLayout, QMessageBox, QDialog
from PySide6.QtGui import QStandardItemModel, QStandardItem
from PySide6.QtCore import Qt, QTimer, QTime
from PySide6.QtGui import QShortcut, QKeySequence
from PySide6.QtGui import QIcon






class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setWindowTitle("SHOW TIMER - PCP 2026")
        self.setWindowIcon(QIcon("icon/icon.png"))

        self.update_shows_table() #a la première ouverture, on insere les emmision depuis la base de donnée dans le tableau 
        
        
        self.selected_show_id = None

        ######BARRE DE MENUS######
        self.ui.actionConfig_OSC.triggered.connect(self.change_osc_config) #quand on clique sur le menu de configuration OSC, on affiche la fenetre de configuration OSC
        self.ui.actionConfig_prompter.triggered.connect(self.change_prompter_config) #quand on clique sur le menu de configuration du prompteur, on affiche la fenetre de configuration du prompteur

        ######BOUTON#######
        self.ui.delete_show.clicked.connect(self.delete_show)
        self.ui.add_show.clicked.connect(self.create_show)
        self.ui.modify_show.clicked.connect(self.modify_show)
        self.ui.open_show.clicked.connect(self.open_show)

        ######TABLEAU######
        self.ui.show_table_view.doubleClicked.connect(self.open_show)
        self.ui.show_table_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.ui.show_table_view.customContextMenuRequested.connect(self.show_table_context_menu)

    def show_table_context_menu(self, pos):
        from PySide6.QtWidgets import QMenu
        if not self.ui.show_table_view.indexAt(pos).isValid():
            return
        menu = QMenu(self)
        menu.addAction("Ouvrir", self.open_show)
        menu.addAction("Modifier", self.modify_show)
        menu.addSeparator()
        menu.addAction("Supprimer", self.delete_show)
        menu.exec(self.ui.show_table_view.viewport().mapToGlobal(pos))

    def update_shows_table(self):
        # Supprime le modèle existant avant de reconstruire
        old_model = self.ui.show_table_view.model()
        if old_model is not None:
            old_model.clear()

        data = database.db.GetAllShows()
        
        if not data:
            return

        columns = list(data[0].keys())

        model = QStandardItemModel(len(data), len(columns))
        model.setHorizontalHeaderLabels(columns)

        for row_idx, row_data in enumerate(data):
            for col_idx, col_name in enumerate(columns):
                value = row_data[col_name]
                
                if col_name == "duree" and value is not None:
                    total_sec = value // 1000
                    minutes = total_sec // 60
                    seconds = total_sec % 60
                    value = f"{minutes:02d}:{seconds:02d}"
                
                item = QStandardItem(str(value))
                model.setItem(row_idx, col_idx, item)

                self.ui.show_table_view.setModel(model)

        from PySide6.QtWidgets import QHeaderView
        header = self.ui.show_table_view.horizontalHeader()
        header.setStretchLastSection(False)
        for i in range(model.columnCount()):
            col_name = model.horizontalHeaderItem(i).text()
            if col_name == "description":
                header.setSectionResizeMode(i, QHeaderView.Stretch)
            else:
                header.setSectionResizeMode(i, QHeaderView.ResizeToContents)

        self.ui.show_table_view.selectionModel().selectionChanged.connect(self.on_show_selected) #reconnecte le signal apres MAJ, sinon plus possible de selec un show...

    def on_show_selected(self): #a chaques clique sur une ligne du tableau, l'id de l'emmision est sauvegardé.
        indexes = self.ui.show_table_view.selectedIndexes()
        if not indexes:
            return

        model = self.ui.show_table_view.model()
        selected_row = indexes[0].row()

        col_names = [model.horizontalHeaderItem(i).text() for i in range(model.columnCount())]
        id_col_index = col_names.index("id")

        self.selected_show_id = model.item(selected_row, id_col_index).text()
        print(f"Emmission sélectionné - ID : {self.selected_show_id}")
    
    def delete_show(self): #fonction appelé quand on clique sur le bouton de suppression d'une emmsion.
        if self.selected_show_id is None:
            QMessageBox.warning(self, "Aucune emmsion sélectionnée", "Veuillez sélectionner une emmsion à supprimer.")
            return
        reply = QMessageBox.question(
        self,
        "Tu confirme chef ?",
        f"Voulez-vous vraiment supprimer l'emmission avec l'id {self.selected_show_id} ?",
        QMessageBox.Yes | QMessageBox.No,
        QMessageBox.No
    )

        if reply == QMessageBox.Yes:
            logger.info(f"Émission supprimée (id={self.selected_show_id})")
            database.db.DeleteShow(self.selected_show_id)
            self.update_shows_table() #apres la suppression, on met a jour le tableau pour que la suppression soit visible
        else:
            print("Suppression annulée.")

    def create_show(self): #fonction appelé quand on clique sur le bouton de création d'une emmsion, elle affiche la fenetre de création d'une emmsion, et si la creation est validé, on met a jour le tableau pour que la nouvelle emmsion soit visible.

        dialog = utils.CreateShowDialog()
        self.hide()
        if dialog.exec() == QDialog.Accepted:
            self.update_shows_table()
            self.show()
        else:
            print("Création d'emmission annulée.")
            self.show()

    def modify_show(self): #fonction appelé quand on clique sur le bouton de modification d'une emmsion
        if self.selected_show_id is None:
                QMessageBox.warning(self, "Aucune emmsion sélectionnée", "Veuillez sélectionner une emmsion à modifier.")
                return
        dialog = utils.ModifyShowDialog(self.selected_show_id)
        self.hide()
        if dialog.exec() == QDialog.Accepted:
            self.update_shows_table()
            self.show()
        else:
            print("Modification d'emmission annulée.")
            self.show()
    
    def open_show(self):
        if self.selected_show_id is None:
            QMessageBox.warning(self, "Aucune emmsion sélectionnée", "Veuillez sélectionner une emmsion à ouvrir.")
            return
        
        database.db.SetActiveShow(self.selected_show_id)
        logger.info(f"Émission ouverte (id={self.selected_show_id})")
        self.operator_dialog = OperatorDialog()  
        self.hide()
        self.operator_dialog.show()

    def closeEvent(self, event):
        reply = QMessageBox.question(
            self,
            "Quitter",
            "Voulez-vous vraiment quitter ShowTimer ?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()

    def change_osc_config(self): #interface de changement de la config OSC
        dialog = utils.ModifyOSCConfigDialog()
        #self.hide()
        if dialog.exec() == QDialog.Accepted:
            print("Configuration OSC mise à jour.")
            #self.show()
        else:
            print("Modification de la configuration OSC annulée.")
            #self.show()

    def change_prompter_config(self): #interface de changement des couleurs, timer, etc du prompteur
        dialog = PrompterSettingsDialog(parent=self)
        dialog.exec()

class OperatorDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_OperatorDialog()
        self.setWindowFlags(Qt.Window)
        self.ui.setupUi(self)
        self.setWindowTitle("Interface opérateur - SHOW TIMER - PCP 2026")
        self.setWindowIcon(QIcon("icon/icon.png"))
        self.setModal(False)
        
        index = self.ui.tabWidget.indexOf(self.ui.fichier)
        if index >= 0:
            self.ui.tabWidget.setCurrentIndex(index)

        self.duree_totale_ms = self.get_total_time()  #duree totale de l'emmision
        self.derniere_seconde_affichee = -1 #derniere seconde affichée
        self.last_next_cues = []
        self._table_next_row    = -1
        self._table_blink_state = False
        self._table_blink_mode  = None  # None | "first" | "second" | "third"

        self.load_screens()

        self.config_table() #config le tableau des cues
        self.config_timer() #confiog le timer, a partir du temps global de l'emmision
        self.fill_table() #remplit le tableau des cues
        self.update_buttons_state("stopped")  # état initial
        self.config_timeline()

        self.load_osc_active_in_ui() #met a jour la case a cocher de l'osc et les deux label 
        self.load_data_in_ui() #affiche le nom, la desc et le temp du show dans l'ui

        

        ########LOGS########
        self.ui.open_logs.clicked.connect(self._open_logs)
        self.ui.delete_logs.clicked.connect(self._delete_logs)
        self._refresh_logs_text()
        logger.set_ui_callback(self._append_log)

        ########BARRE DE FONCTIONS - MENU FICHIER########
        self.ui.quit.clicked.connect(self.close)
        self.ui.lock.clicked.connect(self.lock_interface)
        self.ui.excel_import.clicked.connect(lambda: self.import_from_file("excel"))
        self.ui.csv_import.clicked.connect(lambda: self.import_from_file("csv"))
        self.ui.excel_export.clicked.connect(lambda: self.export_cues("excel"))
        self.ui.csv_export.clicked.connect(lambda: self.export_cues("csv"))
        self.ui.pdf_export.clicked.connect(self._export_pdf)


        ############BARRE DE FONCTIONS - MENU OSC########
        self.ui.osc_active.toggled.connect(self.change_osc_activ_state) #activer / desactiver l'osc
        self.ui.osc_config.clicked.connect(self.change_osc_config) #quand on clique sur le menu de configuration OSC, on affiche la fenetre de configuration OSC, quand fermé, mise a jour de l'ui
        self.ui.osc_ping.clicked.connect(self.test_osc)


        self.ui.cue_table_widget.setStyleSheet("""
            QTableWidget::item:selected {
                background-color: #0078d4;
                color: white;
            }
        """)



        self.prompteur_window = PrompterWindow(self.timer, self.duree_totale_ms) #fenetre du prompteur
        self.ui.screen_active.toggled.connect(self.checkbox_screen_active) #activer / desactiver l'affichage du prompteur
        self.ui.screen_selection.currentIndexChanged.connect(self.screen_selection_changed) #changer l'ecran d'affichage du prompteur 
        self.ui.fullscreen.toggled.connect(self.fullscreen_state_changed) #activer / desactiver le plein ecran du prompteur

        self.ui.prompter_config.clicked.connect(self.config_prompter)

        self.ui.prompt_textedit.setPlaceholderText("Saisissez ici un texte à afficher sur le prompteur...")
        self.ui.prompt_textedit.textChanged.connect(
            lambda: self.prompteur_window.set_prompt_text(self.ui.prompt_textedit.toPlainText())
        )

        #choix de la couleur de fond = mode chroma avec vert et mode luma avec noir
        self.ui.key_mode.addItem("Chroma key (fond vert)", userData="#00FF00")
        self.ui.key_mode.addItem("Luma key (fond noir)",   userData="#000000")

        self.ui.key_mode.currentIndexChanged.connect(self.key_mode_changed)

        #######bouton d'envoie de texte au prompteur########
        self.ui.prompt_highlight.clicked.connect(lambda: self.prompteur_window.start_blink())
        self.ui.prompt_clear.clicked.connect(lambda: self.ui.prompt_textedit.clear())

        self.ui.prompt_highlight_2.clicked.connect(lambda: self.prompteur_window.start_blink())
        self.ui.prompt_clear_2.clicked.connect(lambda: self.ui.prompt_textedit.clear())

        #########Raccourcis clavier##########
        QShortcut(QKeySequence("F1"), self).activated.connect(self.ui.prompt_textedit.clear)  # Efface le texte du prompteur
        QShortcut(QKeySequence("F2"), self).activated.connect(lambda: self.prompteur_window.start_blink())  # Fait clignoter le prompteur

        self._connect_prompter_window_signals()


        ########BOUTONS GESTION DES CUES#############
        self.ui.add_cue.clicked.connect(lambda: self.add_cue_function())
        self.ui.modify_cue.clicked.connect(lambda: self.modify_cue_function())
        self.ui.delete_cue.clicked.connect(lambda: self.delete_cue_function())
        self.ui.delete_all_cues.clicked.connect(lambda: self.delete_all_cues_function())

        ########ZOOM TABLEAU########
        self.ui.zoom_spin.setMinimum(50)
        self.ui.zoom_spin.setMaximum(200)
        self.ui.zoom_spin.setValue(100)
        self.ui.zoom_spin.setSuffix(" %")
        self.ui.zoom_spin.setSingleStep(5)
        self.ui.zoom.clicked.connect(self.zoom_in_table)
        self.ui.unzoom.clicked.connect(self.zoom_out_table)
        self.ui.zoom_spin.valueChanged.connect(self.on_zoom_changed)

    ##################INITIALISATION DES DONNEES DU SHOW#######################
    def load_data_in_ui(self):
        try:
            data = database.db.GetShowById(database.db.GetActiveShow())

            self.ui.show_title.setText(data["nom"])
            self.ui.show_desc.setText(data["description"])
            
            total_ms = data["duree"]
            total_sec = total_ms // 1000
            minutes = total_sec // 60
            seconds = total_sec % 60
            self.ui.show_total_time.setTime(QTime(0, minutes, seconds))


        except Exception as e:
            logger.error(f"Impossible de charger les données du show dans l'interface : {e}")
            QMessageBox.critical(self, "Erreur", "Impossible de charger les données de l'emmsion dans l'interface.")

    ###################PARAMETRE DU PROMPTEUR########################################
    def config_prompter(self):
        if self.prompteur_window.isVisible():
            reply = QMessageBox.question(
                self,
                "Paramètres prompteur",
                "La fenêtre du prompteur va être fermée puis relancée pour appliquer les modifications.\n\nContinuer ?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.No:
                return

        was_visible = self.prompteur_window.isVisible()
        was_fullscreen = self.ui.fullscreen.isChecked()
        was_geometry = self.prompteur_window.geometry() if not was_fullscreen else None
        screen_index = self.ui.screen_selection.currentIndex()
        key_color = self.ui.key_mode.currentData()
        if PrompterSettingsDialog(self).exec_():
            if self.prompteur_window:
                self.prompteur_window.close()
                self.prompteur_window = PrompterWindow(self.timer, self.duree_totale_ms)
                self._connect_prompter_window_signals()
                if was_visible:
                    self.prompteur_window.set_background_color(key_color)
                    if was_fullscreen:
                        self.prompteur_window.set_screen(self.ui.screen_selection.itemData(screen_index))
                        self.prompteur_window.set_fullscreen(True)
                    else:
                        self.prompteur_window.setGeometry(was_geometry)
                        self.prompteur_window.show()
                    self.ui.screen_active.setChecked(True)

    def _connect_prompter_window_signals(self):
        self.prompteur_window.closed_window.connect(
            lambda: self.ui.screen_active.setChecked(False)
        )
        self.prompteur_window.fullscreen_changed.connect(
            lambda checked: self.ui.fullscreen.setChecked(checked)
        )

    def checkbox_screen_active(self, checked):
        if checked:
            self.prompteur_window.show()
        else:
            self.prompteur_window.hide()

    def screen_selection_changed(self, index):
        self.prompteur_window.set_screen(self.ui.screen_selection.currentData())

    def fullscreen_state_changed(self, checked):
        self.prompteur_window.set_fullscreen(checked)

    def key_mode_changed(self, index):
        color = self.ui.key_mode.currentData()
        self.prompteur_window.set_background_color(color)


    ##################ACTIVATION OU NON DE L'ENVOIE OSC#################################
    def change_osc_activ_state(self, state):
        state = self.ui.osc_active.isChecked()

        if state == True:
            database.db.SetOSCState("True")
        else:
            database.db.SetOSCState("False")

    def load_osc_active_in_ui(self):
        state = database.db.GetOSCState()
        if state == "True":
            state = True
            self.ui.osc_active.setChecked(state)
        else:
            state = False
            self.ui.osc_active.setChecked(state)

        osc_ip, osc_port =  database.db.GetOSCConfig()
        
        self.ui.osc_server_ip.setText(osc_ip)
        self.ui.osc_server_port.setValue(int(osc_port))

    def change_osc_config(self):
        if utils.ModifyOSCConfigDialog().exec():
            self.load_osc_active_in_ui()
    

    def test_osc(self):
        try:
            print("test_osc")
            osc.osc_client.send("/test", 1)
            QMessageBox.information(self, "Test OSC", "Message /test avec l'argument int 1 envoyé au serveur OSC")

        except Exception as e:
            logger.error(f"Erreur test OSC : {e}")
            QMessageBox.critical(self, "Test OSC", f"Une erreur s'est produite lors du test OSC : {str(e)}")







    def cue_table_context_menu(self, pos):
        from PySide6.QtWidgets import QMenu
        menu = QMenu(self)
        menu.addAction("Ajouter un cue", self.add_cue_function)

        if self.ui.cue_table_widget.indexAt(pos).isValid() and self.selected_cue_id is not None:
            menu.addSeparator()
            menu.addAction("Modifier", self.modify_cue_function)
            menu.addSeparator()
            menu.addAction("Déclencher OSC manuellement", self.fire_cue_osc_manually)
            menu.addSeparator()
            menu.addAction("Supprimer", self.delete_cue_function)

        menu.exec(self.ui.cue_table_widget.viewport().mapToGlobal(pos))

    def fire_cue_osc_manually(self):
        if self.selected_cue_id is None:
            return
        cue = database.db.GetCueById(self.selected_cue_id)
        if not cue:
            return
        osc_url = cue.get("osc_url", "")
        osc_args = cue.get("osc_args", "")
        if not osc_url:
            QMessageBox.warning(self, "Pas de commande OSC", "Ce cue n'a pas de commande OSC définie.")
            return
        try:
            osc.osc_client.send(osc_url, osc_args)
            QMessageBox.information(self, "OSC envoyé", f"Commande envoyée :\n{osc_url}  {osc_args}")
        except Exception as e:
            logger.error(f"Erreur envoi OSC : {e}")
            QMessageBox.critical(self, "Erreur OSC", f"Impossible d'envoyer la commande OSC :\n{e}")

    #####################################################################################
    def add_cue_function(self):
        if utils.AddModifyCueDialog("ajouter").exec():
            cues = database.db.GetAllCuesFromShow(database.db.GetActiveShow()) or []
            self.timer.set_cues(cues)
            self.fill_table()  # rafraîchit le tableau après l'ajout d'un cue

    def modify_cue_function(self):
        if self.selected_cue_id is None:
            QMessageBox.warning(self, "Aucun cue sélectionné", "Veuillez sélectionner un cue à modifier.")
            return

        dialog = utils.AddModifyCueDialog("modifier", cue_id=self.selected_cue_id)
        if dialog.exec():
            cues = database.db.GetAllCuesFromShow(database.db.GetActiveShow()) or []
            self.timer.set_cues(cues)
            self.fill_table()  # rafraîchit le tableau après la modification d'un cue

    def delete_cue_function(self):
        if self.selected_cue_id is None:
            QMessageBox.warning(self, "Aucun cue sélectionné", "Veuillez sélectionner un cue à supprimer.")
            return

        reply = QMessageBox.question(
            self,
            "Confirmer la suppression",
            f"Voulez-vous vraiment supprimer le cue avec l'id {self.selected_cue_id} ?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            logger.info(f"Cue supprimé (id={self.selected_cue_id})")
            database.db.DeleteCue(self.selected_cue_id)
            cues = database.db.GetAllCuesFromShow(database.db.GetActiveShow()) or []
            self.timer.set_cues(cues)
            self.fill_table()  # rafraîchit le tableau après la suppression d'un cue
        else:
            print("Suppression annulée.")

    def delete_all_cues_function(self):
        try:
            reply = QMessageBox.question(
                self,
                "Confirmer la suppression",
                "Voulez-vous vraiment supprimer tous les cues de l'emission actuelle ?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                logger.info("Tous les cues du show actif supprimés.")
                database.db.DeleteAllCuesFromShow(database.db.GetActiveShow())
                self.fill_table()  # rafraîchit le tableau apres la suppression de tous les cues
            else:
                print("Suppression annulée.")
        except Exception as e:
            logger.error(f"Erreur suppression de tous les cues : {e}")
            QMessageBox.critical(self, "Erreur", f"Une erreur s'est produite lors de la suppression de tous les cues : {str(e)}")

    
    def update_buttons_state(self, etat):
        """
        Gère l'état des boutons selon la situation du timer.
        etat possible : "stopped", "running", "paused"
        """
        if etat == "stopped":
            self.ui.timer_play.setEnabled(True)
            self.ui.timer_pause.setEnabled(False)
            self.ui.timer_stop.setEnabled(False)

        elif etat == "running":
            self.ui.timer_play.setEnabled(False)
            self.ui.timer_pause.setEnabled(True)
            self.ui.timer_stop.setEnabled(True)

        elif etat == "paused":
            self.ui.timer_play.setEnabled(True)   # permet de reprendre
            self.ui.timer_pause.setEnabled(False)
            self.ui.timer_stop.setEnabled(True)

    def config_timeline(self):
        self.timeline = TimelineWidget(self.duree_totale_ms, self.ui.TIMELINE)
        self.timeline.setGeometry(self.ui.TIMELINE.rect())
        self.timer.updated_time.connect(self.timeline.set_position)
        cues = database.db.GetAllCuesFromShow(database.db.GetActiveShow()) or []
        self.timeline.set_cues(cues)

    def zoom_in_table(self):
        val = self.ui.zoom_spin.value()
        self.ui.zoom_spin.setValue(min(val + 5, 200))

    def zoom_out_table(self):
        val = self.ui.zoom_spin.value()
        self.ui.zoom_spin.setValue(max(val - 5, 50))

    def on_zoom_changed(self, percent):
        base_font_size = 11
        base_row_height = 26
        new_size = max(6, round(base_font_size * percent / 100))
        new_row_h = round(base_row_height * percent / 100)

        font = self.ui.cue_table_widget.font()
        font.setPointSize(new_size)
        self.ui.cue_table_widget.setFont(font)
        self.ui.cue_table_widget.horizontalHeader().setFont(font)
        self.ui.cue_table_widget.verticalHeader().setDefaultSectionSize(new_row_h)

        # Applique la police sur chaque cellule existante
        for row in range(self.ui.cue_table_widget.rowCount()):
            for col in range(self.ui.cue_table_widget.columnCount()):
                item = self.ui.cue_table_widget.item(row, col)
                if item:
                    item.setFont(font)

    def config_table(self):
        from PySide6.QtWidgets import QTableWidget, QHeaderView
        self.ui.cue_table_widget.setColumnCount(6)
        self.ui.cue_table_widget.setHorizontalHeaderLabels(["Nom", "Description", "Déclenchement", "URL OSC", "Args OSC", "Dans"])
        self.ui.cue_table_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.ui.cue_table_widget.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.ui.cue_table_widget.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.ui.cue_table_widget.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.ui.cue_table_widget.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.ui.cue_table_widget.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.ui.cue_table_widget.setEditTriggers(QTableWidget.NoEditTriggers)
        self.ui.cue_table_widget.setSelectionBehavior(QTableWidget.SelectRows)

        self.selected_cue_id = None
        self.ui.cue_table_widget.itemSelectionChanged.connect(self.on_cue_selected)
        self.ui.cue_table_widget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.ui.cue_table_widget.customContextMenuRequested.connect(self.cue_table_context_menu)

    def config_timer(self):
        from timer import PrompteurTimer
        self.timer = PrompteurTimer(self)

        self.ui.progressBar.setMinimum(0)
        self.ui.progressBar.setMaximum(self.duree_totale_ms)
        self.ui.progressBar.setValue(0)

        show_id = database.db.GetActiveShow()
        if show_id is not None:
            cues = database.db.GetAllCuesFromShow(show_id) or []
            self.timer.set_cues(cues)

        self.timer.updated_time.connect(self.on_timer_updated)
        self.timer.two_next_cues.connect(self.on_next_cue)
        self.timer.fired_cue.connect(self.on_cue_fired)

        self._table_blink_timer = QTimer(self)
        self._table_blink_timer.timeout.connect(self._on_table_blink_tick)

        self.ui.timer_play.clicked.connect(self.on_play_clicked)
        self.ui.timer_pause.clicked.connect(self.on_pause_clicked)
        self.ui.timer_stop.clicked.connect(self.on_stop_clicked)

    def fill_table(self):
        from PySide6.QtWidgets import QTableWidgetItem
        from PySide6.QtGui import QColor

        self.ui.cue_table_widget.setRowCount(0)

        show_id = database.db.GetActiveShow()
        if show_id is None:
            return

        cues = database.db.GetAllCuesFromShow(show_id) or []

        if hasattr(self, "timeline"):
            self.timeline.set_cues(cues)

        for cue in sorted(cues, key=lambda c: c["temps"]):
            ligne = self.ui.cue_table_widget.rowCount()
            self.ui.cue_table_widget.insertRow(ligne)

            h = cue["temps"] // 3600000
            m = (cue["temps"] % 3600000) // 60000
            s = (cue["temps"] % 60000) // 1000

            self.ui.cue_table_widget.setItem(ligne, 0, QTableWidgetItem(cue["nom"]))
            self.ui.cue_table_widget.setItem(ligne, 1, QTableWidgetItem(cue.get("description", "")))
            self.ui.cue_table_widget.setItem(ligne, 2, QTableWidgetItem(f"{h:02d}:{m:02d}:{s:02d}"))
            self.ui.cue_table_widget.setItem(ligne, 3, QTableWidgetItem(cue.get("osc_url", "") or ""))
            self.ui.cue_table_widget.setItem(ligne, 4, QTableWidgetItem(str(cue.get("osc_args", "") or "")))
            self.ui.cue_table_widget.setItem(ligne, 5, QTableWidgetItem("--:--:--"))

            self.ui.cue_table_widget.item(ligne, 0).setData(Qt.UserRole,     cue["temps"])
            self.ui.cue_table_widget.item(ligne, 0).setData(Qt.UserRole + 1, cue["cue_id"])

            original_bg = QColor(cue["color"]) if cue.get("color") else QColor()
            for colonne in range(self.ui.cue_table_widget.columnCount()):
                item = self.ui.cue_table_widget.item(ligne, colonne)
                item.setData(Qt.UserRole + 2, original_bg)
                if cue.get("color"):
                    item.setBackground(original_bg)

            #les bouton sont par defaut desactivé
            self.ui.modify_cue.setEnabled(False)
            self.ui.delete_cue.setEnabled(False)

    def on_cue_selected(self):
        ligne = self.ui.cue_table_widget.currentRow()
        if ligne < 0:
            self.selected_cue_id = None
            return
        item_nom = self.ui.cue_table_widget.item(ligne, 0)
        if item_nom:
            self.selected_cue_id = item_nom.data(Qt.UserRole + 1)
            print(f"Cue sélectionné - ID : {self.selected_cue_id}")

            # Basculer sur l'onglet Cues
            index = self.ui.tabWidget.indexOf(self.ui.tab)
            if index >= 0:
                self.ui.tabWidget.setCurrentIndex(index)

            #quand un cue est sélectionné, on active les boutons de modification et de suppression de cue
            self.ui.modify_cue.setEnabled(True)
            self.ui.delete_cue.setEnabled(True)

    ###################################BOUTON###############################################3

    def on_play_clicked(self):
        if self.timer.running:
            return

        if self.timer.timeline_time > 0:
            self.timer.resume()
            logger.info("Timer repris.")
        else:
            self.last_next_cues = []
            self.fill_table()
            self.timer.start(0)
            logger.info("Timer démarré.")

        self.update_buttons_state("running")

    def on_pause_clicked(self):
        """Demande une confirmation avant de mettre en pause."""
        if not self.timer.running:
            return

        reponse = QMessageBox.question(
            self,
            "Mettre en pause ?",
            "Voulez-vous mettre le timer en pause ?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reponse == QMessageBox.Yes:
            self.timer.pause()
            self.update_buttons_state("paused")
            logger.info("Timer mis en pause.")

    def on_stop_clicked(self):
        reponse = QMessageBox.question(
                self,
                "Arrêter le timer ?",
                "Voulez-vous arrêter et remettre le timer à zéro ?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

        if reponse == QMessageBox.No:
            return

        self.timer.stop()
        logger.info("Timer arrêté.")
        self.update_buttons_state("stopped")
        self.derniere_seconde_affichee = -1
        self.ui.pass_time.setText("00:00:00")
        self.ui.remaining_time.setText("00:00:00")
        self.ui.remaining_time.setStyleSheet("")
        self.ui.progressBar.setValue(0)
        self.prompteur_window._reset_affichage()

        show_id = database.db.GetActiveShow()
        if show_id is not None:
            cues = database.db.GetAllCuesFromShow(show_id) or []
            self.timer.set_cues(cues)  # ← recharge les cues après le stop

        self.fill_table()

    #######################TIMER############################################3

    def on_timer_updated(self, temps_ms):
        from PySide6.QtGui import QColor

        seconde_actuelle = temps_ms // 1000

        self.ui.progressBar.setValue(temps_ms)

        if seconde_actuelle == self.derniere_seconde_affichee:
            return
        self.derniere_seconde_affichee = seconde_actuelle

        h = temps_ms // 3600000
        m = (temps_ms % 3600000) // 60000
        s = (temps_ms % 60000) // 1000
        self.ui.pass_time.setText(f"{h:02d}:{m:02d}:{s:02d}")

        restant = max(0, self.duree_totale_ms - temps_ms)
        h = restant // 3600000
        m = (restant % 3600000) // 60000
        s = (restant % 60000) // 1000
        self.ui.remaining_time.setText(f"{h:02d}:{m:02d}:{s:02d}")

        if restant < 60_000:
            self.ui.remaining_time.setStyleSheet("color: red;")
        else:
            self.ui.remaining_time.setStyleSheet("")

        if restant == 0:
            self.timer.stop()
            self.update_buttons_state("stopped")

        self.ui.cue_table_widget.setUpdatesEnabled(False)

        for ligne in range(self.ui.cue_table_widget.rowCount()):
            item_nom = self.ui.cue_table_widget.item(ligne, 0)
            if item_nom is None:
                continue
            temps_cue = item_nom.data(Qt.UserRole)
            dans_ms   = temps_cue - temps_ms
            if dans_ms <= 0:
                self.ui.cue_table_widget.item(ligne, 5).setText("passé")
            else:
                h = dans_ms // 3600000
                m = (dans_ms % 3600000) // 60000
                s = (dans_ms % 60000) // 1000
                self.ui.cue_table_widget.item(ligne, 5).setText(f"{h:02d}:{m:02d}:{s:02d}")

        self.ui.cue_table_widget.setUpdatesEnabled(True)

        # Clignotement du prochain cue (seuils fixes)
        if self._table_next_row >= 0 and self.last_next_cues:
            temps_dans = self.last_next_cues[0]["temps"] - temps_ms
            if temps_dans <= 0:
                new_mode, interval = None, None
            elif temps_dans <= 5_000:
                new_mode, interval = "third", 250    # rouge,  rapide
            elif temps_dans <= 10_000:
                new_mode, interval = "second", 450   # orange, moyen
            elif temps_dans <= 15_000:
                new_mode, interval = "first", 800    # vert,   lent
            else:
                new_mode, interval = None, None

            if new_mode != self._table_blink_mode:
                self._table_blink_mode  = new_mode
                self._table_blink_state = False
                self._table_blink_timer.stop()
                if new_mode is not None:
                    self._table_blink_timer.setInterval(interval)
                    self._table_blink_timer.start()
                else:
                    # Remettre le fond vert fixe
                    self._set_row_background(self._table_next_row, QColor("#1a7a1a"))

    def on_cue_fired(self, cue):
        from PySide6.QtGui import QColor

        for ligne in range(self.ui.cue_table_widget.rowCount()):
            item_nom = self.ui.cue_table_widget.item(ligne, 0)
            if item_nom and item_nom.data(Qt.UserRole) == cue["temps"]:
                # On grise définitivement le fond de ce cue
                for colonne in range(self.ui.cue_table_widget.columnCount()):
                    item = self.ui.cue_table_widget.item(ligne, colonne)
                    if item:
                        item.setBackground(QColor("gray"))
                        item.setForeground(QColor("white"))
                self.ui.cue_table_widget.scrollToItem(item_nom)
                break

        logger.info(f"CUE déclenché : '{cue['nom']}' à {cue['temps']}ms  |  OSC: {cue.get('osc_url','')} {cue.get('osc_args','')}")

    def on_next_cue(self, cues):
        from PySide6.QtGui import QColor

        new_ids = [c["cue_id"] for c in cues]
        old_ids = [c["cue_id"] for c in self.last_next_cues]
        if new_ids == old_ids:
            return

        self.last_next_cues = cues

        # Stopper le clignotement — sera redémarré si besoin dans on_timer_updated
        self._table_blink_timer.stop()
        self._table_blink_mode  = None
        self._table_blink_state = False
        self._table_next_row    = -1

        temps_actuel = self.timer.actual_time()
        next_temps   = {c["temps"]: i for i, c in enumerate(cues)}  # temps → index dans cues

        COLOR_PAST        = QColor("#555555")
        COLOR_NEXT        = QColor("#1a7a1a")   # vert
        COLOR_SECOND_NEXT = QColor("#0f4f0f")   # vert foncé
        FG_PAST           = QColor("#999999")
        FG_DEFAULT        = QColor("white")

        self.ui.cue_table_widget.setUpdatesEnabled(False)

        for ligne in range(self.ui.cue_table_widget.rowCount()):
            item_nom = self.ui.cue_table_widget.item(ligne, 0)
            if item_nom is None:
                continue
            temps_cue = item_nom.data(Qt.UserRole)

            if temps_cue <= temps_actuel:
                # Cue passé → fond gris
                bg = COLOR_PAST
                fg = FG_PAST
            elif temps_cue in next_temps:
                idx = next_temps[temps_cue]
                if idx == 0:
                    bg = COLOR_NEXT
                    fg = FG_DEFAULT
                    self._table_next_row = ligne
                else:
                    bg = COLOR_SECOND_NEXT
                    fg = FG_DEFAULT
            else:
                # Cue futur neutre → couleur originale
                orig = item_nom.data(Qt.UserRole + 2)
                bg = orig if (orig and orig.isValid()) else QColor()
                fg = FG_DEFAULT

            for colonne in range(self.ui.cue_table_widget.columnCount()):
                item = self.ui.cue_table_widget.item(ligne, colonne)
                if item:
                    if bg.isValid():
                        item.setBackground(bg)
                    else:
                        item.setData(Qt.BackgroundRole, None)
                    item.setForeground(fg)

        self.ui.cue_table_widget.setUpdatesEnabled(True)

    def _set_row_background(self, ligne, color):
        for colonne in range(self.ui.cue_table_widget.columnCount()):
            item = self.ui.cue_table_widget.item(ligne, colonne)
            if item:
                if color is None or not color.isValid():
                    item.setData(Qt.BackgroundRole, None)
                else:
                    item.setBackground(color)

    def _on_table_blink_tick(self):
        from PySide6.QtGui import QColor
        if self._table_next_row < 0:
            return
        self._table_blink_state = not self._table_blink_state
        if self._table_blink_state:
            # Couleur vive selon le mode
            colors = {
                "first":  QColor("#00CC00"),   # vert   — ≤ 15s
                "second": QColor("#FF8800"),   # orange — ≤ 10s
                "third":  QColor("#FF0000"),   # rouge  — ≤ 5s
            }
            self._set_row_background(self._table_next_row, colors.get(self._table_blink_mode, QColor("#00DD00")))
        else:
            self._set_row_background(self._table_next_row, QColor("#1a7a1a"))

    def get_total_time(self):
        show_id = database.db.GetActiveShow()
        if show_id is None:
            return 0
        show = database.db.GetShowById(show_id)
        return show["duree"] if show else 0
    
    def load_screens(self): #charge les ecran disponibles

        ecrans = QApplication.screens()
        
        self.ui.screen_selection.clear()
        
        for i, ecran in enumerate(ecrans):
            nom = ecran.name()
            resolution = ecran.geometry()
            self.ui.screen_selection.addItem(
                f"Écran {i} - {nom} ({resolution.width()}x{resolution.height()})",
                userData=i  # on stocke le numéro de l'écran dans le userData
            )

    def import_from_file(self, mode):
        dialog = utils.ImportExcelDialog(self, mode=mode)
        if mode == "excel":
            dialog._browse_filter = "Excel (*.xlsx *.xls)"
        else:
            dialog._browse_filter = "CSV (*.csv)"
        if dialog.exec():
            show_id = database.db.GetActiveShow()
            for cue in dialog.cues:
                database.db.AddCueToShow(
                    show_id,
                    cue["nom"],
                    cue["description"],
                    cue["temps"],
                    cue["osc_url"],
                    cue["osc_args"],
                    "#161A16"
                )
            cues = database.db.GetAllCuesFromShow(show_id) or []
            self.timer.set_cues(cues)
            self.fill_table()
            logger.info(f"{len(dialog.cues)} cue(s) importés depuis fichier ({mode.upper()}).")
            QMessageBox.information(self, "Import réussi", f"{len(dialog.cues)} cue(s) importé(s).")

    def export_cues(self, mode):
        from PySide6.QtWidgets import QFileDialog, QMessageBox
        from datetime import datetime  # Ajout pour la date et l'heure

        show_id = database.db.GetActiveShow()
        if show_id is None:
            QMessageBox.warning(self, "Aucun show", "Aucun show actif.")
            return

        cues = database.db.GetAllCuesFromShow(show_id) or []
        if not cues:
            QMessageBox.warning(self, "Aucun cue", "Aucun cue à exporter.")
            return

        # Génération du suffixe date_heure et nom du show
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        show = database.db.GetShowById(show_id)
        show_name = show.get("nom", "show") if show else "show"
        show_name = "".join(c for c in show_name if c.isalnum() or c in " _-").strip().replace(" ", "_")
        
        def ms_to_mmss(ms):
            m = ms // 60000
            s = (ms % 60000) // 1000
            return f"{m:02d}:{s:02d}"

        rows = [["TITRE", "DESCRIPTION", "TEMPS", "URL_OSC", "ARG_OSC"]]
        for cue in sorted(cues, key=lambda c: c["temps"]):
            rows.append([
                cue.get("nom", ""),
                cue.get("description", ""),
                ms_to_mmss(cue.get("temps", 0)),
                cue.get("osc_url", ""),
                str(cue.get("osc_args", "") or ""),
            ])

        if mode == "excel":
            default_name = f"{show_name}_cues_{timestamp}.xlsx"
            path, _ = QFileDialog.getSaveFileName(self, "Exporter en Excel", default_name, "Excel (*.xlsx)")
            
            if not path:
                return
                
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cues"
            for i, row in enumerate(rows):
                ws.append(row)
                if i == 0:
                    for cell in ws[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill("solid", fgColor="0078D4")
            
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 40
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 30
            ws.column_dimensions["E"].width = 12
            wb.save(path)

        elif mode == "csv":
            default_name = f"{show_name}_cues_{timestamp}.csv"
            path, _ = QFileDialog.getSaveFileName(self, "Exporter en CSV", default_name, "CSV (*.csv)")
            
            if not path:
                return
                
            import csv
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f, delimiter=";")
                writer.writerows(rows)

        QMessageBox.information(self, "Export réussi", f"{len(cues)} cue(s) exporté(s) vers :\n{path}")

    def _export_pdf(self):
        from PySide6.QtWidgets import QFileDialog, QMessageBox
        from datetime import datetime

        show_id = database.db.GetActiveShow()
        if show_id is None:
            QMessageBox.warning(self, "Aucun show", "Aucun show actif.")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        show = database.db.GetShowById(show_id)
        show_name = show.get("nom", "show") if show else "show"
        show_name = "".join(c for c in show_name if c.isalnum() or c in " _-").strip().replace(" ", "_")
        default_name = f"{show_name}_rapport_{timestamp}.pdf"
        path, _ = QFileDialog.getSaveFileName(self, "Exporter en PDF", default_name, "PDF (*.pdf)")
        if not path:
            return

        try:
            export_show_to_pdf(show_id, path)
            QMessageBox.information(self, "Export réussi", f"PDF exporté vers :\n{path}")
        except Exception as e:
            logger.error(f"Erreur export PDF : {e}")
            QMessageBox.critical(self, "Erreur export PDF", str(e))

    # ── Logs ──────────────────────────────────────────────────
    def _append_log(self, line):
        """Appelé par logger à chaque nouvelle entrée — mise à jour en temps réel."""
        self.ui.logs_text.append(line)
        self.ui.logs_text.verticalScrollBar().setValue(
            self.ui.logs_text.verticalScrollBar().maximum()
        )

    def closeEvent(self, event):
        logger.clear_ui_callback()
        event.accept()

    def _refresh_logs_text(self):
        """Affiche les dernières lignes du log courant dans logs_text."""
        path = logger.get_log_file_path()
        if path and os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
            self.ui.logs_text.setPlainText(content)
            self.ui.logs_text.verticalScrollBar().setValue(
                self.ui.logs_text.verticalScrollBar().maximum()
            )
        else:
            self.ui.logs_text.setPlainText("Aucun log disponible.")

    def _open_logs(self):
        """Ouvre le dossier des logs dans l'explorateur."""
        logs_dir = logger.get_logs_dir()
        os.makedirs(logs_dir, exist_ok=True)
        os.startfile(logs_dir)

    def _delete_logs(self):
        """Supprime tous les fichiers de logs après confirmation."""
        logs_dir = logger.get_logs_dir()
        if not os.path.exists(logs_dir):
            QMessageBox.information(self, "Logs", "Aucun fichier de log à supprimer.")
            return

        fichiers = [f for f in os.listdir(logs_dir) if f.endswith(".txt")]
        if not fichiers:
            QMessageBox.information(self, "Logs", "Aucun fichier de log à supprimer.")
            return

        reply = QMessageBox.question(
            self,
            "Supprimer les logs",
            f"Voulez-vous vraiment supprimer {len(fichiers)} fichier(s) de log ?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            for f in fichiers:
                try:
                    os.remove(os.path.join(logs_dir, f))
                except Exception:
                    pass
            logger.info("Fichiers de logs supprimés par l'utilisateur.")
            self._refresh_logs_text()
            QMessageBox.information(self, "Logs", "Fichiers de logs supprimés.")

    def lock_interface(self):
        from PySide6.QtWidgets import QWidget
        from PySide6.QtCore import Qt

        # Overlay gris semi-transparent par dessus la fenêtre
        overlay = QWidget(self)
        overlay.setGeometry(self.rect())
        overlay.setStyleSheet("background-color: rgba(0, 0, 0, 160);")
        overlay.setAttribute(Qt.WA_TransparentForMouseEvents, False)
        overlay.show()
        overlay.raise_()

        dialog = utils.LockDialog(self)
        # Centrer sur la fenêtre opérateur
        geo = self.geometry()
        dx = geo.x() + (geo.width() - dialog.width()) // 2
        dy = geo.y() + (geo.height() - dialog.height()) // 2
        dialog.move(dx, dy)
        dialog.exec()

        overlay.deleteLater()

    def closeEvent(self, event):
        reponse = QMessageBox.question(
            self,
            "Fermer ?",
            "Voulez-vous fermer l'interface opérateur ?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reponse == QMessageBox.Yes:
            if self.timer.running:
                self.timer.stop()
            self.prompteur_window.hide()
            from PySide6.QtWidgets import QApplication
            for widget in QApplication.topLevelWidgets():
                if isinstance(widget, MainWindow):
                    widget.show()
                    break
            event.accept()
        else:
            event.ignore()
    

if __name__ == "__main__":
    import os
    if getattr(sys, 'frozen', False):
        os.chdir(sys._MEIPASS)  # ressources dans _internal/ (icônes, etc.)

    logger.setup()
    utils.init_password_file()

    def _handle_exception(exc_type, exc_value, exc_tb):
        import traceback
        logger.error("Exception non gérée : " + "".join(traceback.format_exception(exc_type, exc_value, exc_tb)))
        sys.__excepthook__(exc_type, exc_value, exc_tb)
    sys.excepthook = _handle_exception

    database.connect_db()
    osc.connect_osc_client()

    #database.db.AddCueToShow(1, "Test Cue 2", "Description du cue 2", 130000, "/test/cue2", "arg1 arg2", "#189E35") #ajout de cue de test dans la base de donnée, pour pouvoir tester l'interface de controle et le timer



    



    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon("icon/icon.png"))
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
