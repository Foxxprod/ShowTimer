from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QPushButton, QLabel,
    QLineEdit, QTextEdit, QHBoxLayout, QGridLayout, QMessageBox, QDialog,
    QTimeEdit, QFormLayout, QGroupBox, QSpinBox, QDialogButtonBox, QScrollArea,
    QColorDialog, QSizePolicy, QTableWidget, QTableWidgetItem, QFileDialog, QHeaderView
)
from PySide6.QtGui import QStandardItemModel, QStandardItem, QColor, QPainter, QPainterPath, QFont, QPen, QPixmap
from PySide6.QtCore import QTime, Qt, Signal

import database, osc, logger, sys, os


from ui.ui_new_show import Ui_Dialog
from ui.ui_osc_config import Ui_Dialog as Ui_OSCDialog
from ui.ui_add_modify_cue import Ui_Dialog as Ui_CueDialog

#classe qui contient les label utilisé pour les timer avec la bordure
#On remplace le widget generique par ce widget custom
#C'est pas du texte c'est une image vectorisé
class LabelWithOutline(QWidget):
    def __init__(self, text="", color=None, font_size=36, outline_width=4, parent=None):
        super().__init__(parent)
        self._text = text
        self._color = QColor(color) if isinstance(color, str) else (color or QColor("white"))
        self._font_size = int(font_size)
        self._outline_width = int(outline_width)

    def set_text(self, text):
        self._text = text
        self.update()

    def set_color(self, color):
        self._color = QColor(color) if isinstance(color, str) else color
        self.update()

    def set_font_size(self, size):
        self._font_size = int(size)
        self.update()

    def set_outline_width(self, width):
        self._outline_width = int(width)
        self.update()

    def paintEvent(self, event):
        if not self._text:
            return
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        font = QFont("Arial", self._font_size, QFont.Bold)
        path = QPainterPath()
        path.addText(0, 0, font, self._text)
        text_rect = path.boundingRect()
        if text_rect.width() == 0 or text_rect.height() == 0:
            return
        scale_x = self.width() / text_rect.width() if text_rect.width() > self.width() else 1.0
        scale_y = self.height() / text_rect.height() if text_rect.height() > self.height() else 1.0
        scale = min(scale_x, scale_y)
        painter.translate(self.width() / 2, self.height() / 2)
        painter.scale(scale, scale)
        painter.translate(-text_rect.center())
        painter.setPen(QPen(QColor("black"), self._outline_width / scale))
        painter.setBrush(self._color)
        painter.drawPath(path)


class ColorButton(QPushButton):
    color_changed = Signal(str)

    def __init__(self, color="#FFFFFF", parent=None):
        super().__init__(parent)
        self._color = QColor(color)
        self._refresh()
        self.clicked.connect(self._pick)

    def _refresh(self):
        self.setText(self._color.name().upper())
        fg = "#000000" if self._color.lightness() > 128 else "#FFFFFF"
        self.setStyleSheet(f"""
            QPushButton {{
                background-color: {self._color.name()};
                color: {fg};
                border: 1px solid #888;
                padding: 4px 10px;
                border-radius: 3px;
                font-weight: bold;
            }}
        """)

    def _pick(self):
        c = QColorDialog.getColor(self._color, self, "Choisir une couleur")
        if c.isValid():
            self._color = c
            self._refresh()
            self.color_changed.emit(self._color.name())

    def get_color(self):
        return self._color.name()

    def set_color(self, color_str):
        self._color = QColor(color_str)
        self._refresh()


class PrompterSettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Configuration du Prompteur")
        self.setMinimumSize(950, 620)
        self.resize(1150, 700)

        self.config = database.db.GetPrompterConfig()
        self._build_ui()
        self._load_config()
        self._update_preview()

    def _build_ui(self):
        outer = QVBoxLayout(self)

        content = QHBoxLayout()
        content.setSpacing(12)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedWidth(440)
        scroll.setStyleSheet("QScrollArea { border: none; }")

        form_widget = QWidget()
        form_vbox = QVBoxLayout(form_widget)
        form_vbox.setSpacing(10)
        form_vbox.addWidget(self._group_prompt_text())
        form_vbox.addWidget(self._group_next_cue())
        form_vbox.addWidget(self._group_second_cue())
        form_vbox.addWidget(self._group_clock())
        form_vbox.addWidget(self._group_blink())
        form_vbox.addStretch()

        scroll.setWidget(form_widget)
        content.addWidget(scroll)

        content.addWidget(self._build_preview(), 1)

        outer.addLayout(content)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.button(QDialogButtonBox.Ok).setText("Enregistrer")
        btn_box.button(QDialogButtonBox.Cancel).setText("Annuler")
        btn_box.accepted.connect(self._save)
        btn_box.rejected.connect(self.reject)
        outer.addWidget(btn_box)

    def _make_group(self, title):
        group = QGroupBox(title)
        form = QFormLayout(group)
        form.setLabelAlignment(Qt.AlignRight)
        form.setSpacing(8)
        return group, form

    def _make_spinbox(self, min_val, max_val, suffix=""):
        sb = QSpinBox()
        sb.setRange(min_val, max_val)
        if suffix:
            sb.setSuffix(suffix)
        sb.valueChanged.connect(self._update_preview)
        return sb

    def _group_prompt_text(self):
        group, form = self._make_group("Texte principal")
        self.sb_prompter_text_size = self._make_spinbox(8, 300, " px")
        self.cb_prompter_text_color = ColorButton()
        self.cb_prompter_text_color.color_changed.connect(self._update_preview)
        form.addRow("Taille :", self.sb_prompter_text_size)
        form.addRow("Couleur :", self.cb_prompter_text_color)
        return group

    def _group_next_cue(self):
        group, form = self._make_group("Next Cue")
        self.sb_next_cue_text_size = self._make_spinbox(8, 300, " px")
        self.cb_next_cue_text_color = ColorButton()
        self.cb_next_cue_text_color.color_changed.connect(self._update_preview)
        self.sb_next_cue_outline = self._make_spinbox(0, 20, " px")
        form.addRow("Taille :", self.sb_next_cue_text_size)
        form.addRow("Couleur :", self.cb_next_cue_text_color)
        form.addRow("Contour :", self.sb_next_cue_outline)
        return group

    def _group_second_cue(self):
        group, form = self._make_group("Second Cue")
        self.sb_second_cue_text_size = self._make_spinbox(8, 300, " px")
        self.cb_second_cue_text_color = ColorButton()
        self.cb_second_cue_text_color.color_changed.connect(self._update_preview)
        self.sb_second_cue_outline = self._make_spinbox(0, 20, " px")
        form.addRow("Taille :", self.sb_second_cue_text_size)
        form.addRow("Couleur :", self.cb_second_cue_text_color)
        form.addRow("Contour :", self.sb_second_cue_outline)
        return group

    def _group_clock(self):
        group, form = self._make_group("Horloge")
        self.sb_clock_text_size = self._make_spinbox(8, 300, " px")
        self.cb_clock_text_color = ColorButton()
        self.cb_clock_text_color.color_changed.connect(self._update_preview)
        self.sb_clock_outline = self._make_spinbox(0, 20, " px")
        form.addRow("Taille :", self.sb_clock_text_size)
        form.addRow("Couleur :", self.cb_clock_text_color)
        form.addRow("Contour :", self.sb_clock_outline)
        return group

    def _group_blink(self):
        group, form = self._make_group("Clignotements du Next Cue")

        self.sb_blink_first_time = self._make_spinbox(1, 60, " s")
        self.cb_blink_first_color = ColorButton()
        self.cb_blink_first_color.color_changed.connect(self._update_preview)

        self.sb_blink_second_time = self._make_spinbox(1, 60, " s")
        self.cb_blink_second_color = ColorButton()
        self.cb_blink_second_color.color_changed.connect(self._update_preview)

        self.sb_blink_third_time = self._make_spinbox(1, 60, " s")
        self.cb_blink_third_color = ColorButton()
        self.cb_blink_third_color.color_changed.connect(self._update_preview)

        form.addRow("Seuil 1 (lent) :", self.sb_blink_first_time)
        form.addRow("Couleur 1 :", self.cb_blink_first_color)
        form.addRow("Seuil 2 :", self.sb_blink_second_time)
        form.addRow("Couleur 2 :", self.cb_blink_second_color)
        form.addRow("Seuil 3 (urgent) :", self.sb_blink_third_time)
        form.addRow("Couleur 3 :", self.cb_blink_third_color)
        return group

    def _build_preview(self):
        container = QWidget()
        container.setStyleSheet("background-color: #111111; border-radius: 8px;")

        vbox = QVBoxLayout(container)
        vbox.setContentsMargins(14, 14, 14, 14)
        vbox.setSpacing(6)

        lbl_title = QLabel("Aperçu")
        lbl_title.setStyleSheet("color: #555555; font-size: 11px; background: transparent;")
        lbl_title.setAlignment(Qt.AlignCenter)
        vbox.addWidget(lbl_title)

        self.preview_prompt = QLabel("Texte du prompteur — exemple affiché à l'écran")
        self.preview_prompt.setWordWrap(True)
        self.preview_prompt.setAlignment(Qt.AlignCenter)
        self.preview_prompt.setMinimumHeight(100)
        self.preview_prompt.setStyleSheet("background: transparent;")
        vbox.addWidget(self.preview_prompt)

        vbox.addStretch()

        lbl_s = QLabel("Second Cue")
        lbl_s.setStyleSheet("color: #444444; font-size: 10px; background: transparent;")
        vbox.addWidget(lbl_s)

        self.preview_second_cue = LabelWithOutline("Suite émission : 00:08:30")
        self.preview_second_cue.setMinimumHeight(55)
        vbox.addWidget(self.preview_second_cue)

        lbl_n = QLabel("Next Cue")
        lbl_n.setStyleSheet("color: #444444; font-size: 10px; background: transparent;")
        vbox.addWidget(lbl_n)

        self.preview_next_cue = LabelWithOutline("Plateau 1 : 00:02:15")
        self.preview_next_cue.setMinimumHeight(90)
        vbox.addWidget(self.preview_next_cue)

        lbl_c = QLabel("Horloge")
        lbl_c.setStyleSheet("color: #444444; font-size: 10px; background: transparent;")
        vbox.addWidget(lbl_c)

        self.preview_clock = LabelWithOutline("Écoulé : 00:10:30     Restant : 01:49:30")
        self.preview_clock.setMinimumHeight(55)
        vbox.addWidget(self.preview_clock)

        lbl_blink = QLabel("Couleurs de clignotement")
        lbl_blink.setStyleSheet("color: #444444; font-size: 10px; background: transparent;")
        vbox.addWidget(lbl_blink)

        blink_row = QHBoxLayout()
        blink_row.setSpacing(6)
        self.preview_blink1 = QLabel()
        self.preview_blink2 = QLabel()
        self.preview_blink3 = QLabel()
        for lbl in [self.preview_blink1, self.preview_blink2, self.preview_blink3]:
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setMinimumHeight(32)
            lbl.setStyleSheet("border-radius: 4px; font-weight: bold; font-size: 12px;")
            blink_row.addWidget(lbl)
        vbox.addLayout(blink_row)

        return container

    def _load_config(self):
        c = self.config
        self.sb_prompter_text_size.setValue(int(c["prompter_text_size"]))
        self.cb_prompter_text_color.set_color(c["prompter_text_color"])

        self.sb_next_cue_text_size.setValue(int(c["next_cue_text_size"]))
        self.cb_next_cue_text_color.set_color(c["next_cue_text_color"])
        self.sb_next_cue_outline.setValue(int(c["next_cue_outline_width"]))

        self.sb_second_cue_text_size.setValue(int(c["second_cue_text_size"]))
        self.cb_second_cue_text_color.set_color(c["second_cue_text_color"])
        self.sb_second_cue_outline.setValue(int(c["second_cue_outline_width"]))

        self.sb_clock_text_size.setValue(int(c["clock_text_size"]))
        self.cb_clock_text_color.set_color(c["clock_text_color"])
        self.sb_clock_outline.setValue(int(c["clock_outline_width"]))

        self.sb_blink_first_time.setValue(int(c["blink_first_time"]) // 1000)
        self.cb_blink_first_color.set_color(c["blink_first_color"])
        self.sb_blink_second_time.setValue(int(c["blink_second_time"]) // 1000)
        self.cb_blink_second_color.set_color(c["blink_second_color"])
        self.sb_blink_third_time.setValue(int(c["blink_third_time"]) // 1000)
        self.cb_blink_third_color.set_color(c["blink_third_color"])

    def _update_preview(self):
        size = self.sb_prompter_text_size.value()
        color = self.cb_prompter_text_color.get_color()
        self.preview_prompt.setStyleSheet(f"""
            color: {color};
            font-size: {size}px;
            font-weight: bold;
            background-color: transparent;
        """)

        self.preview_next_cue.set_font_size(self.sb_next_cue_text_size.value())
        self.preview_next_cue.set_color(self.cb_next_cue_text_color.get_color())
        self.preview_next_cue.set_outline_width(self.sb_next_cue_outline.value())

        self.preview_second_cue.set_font_size(self.sb_second_cue_text_size.value())
        self.preview_second_cue.set_color(self.cb_second_cue_text_color.get_color())
        self.preview_second_cue.set_outline_width(self.sb_second_cue_outline.value())

        self.preview_clock.set_font_size(self.sb_clock_text_size.value())
        self.preview_clock.set_color(self.cb_clock_text_color.get_color())
        self.preview_clock.set_outline_width(self.sb_clock_outline.value())

        for lbl, btn, sb in [
            (self.preview_blink1, self.cb_blink_first_color,  self.sb_blink_first_time),
            (self.preview_blink2, self.cb_blink_second_color, self.sb_blink_second_time),
            (self.preview_blink3, self.cb_blink_third_color,  self.sb_blink_third_time),
        ]:
            c = btn.get_color()
            fg = "#000000" if QColor(c).lightness() > 128 else "#FFFFFF"
            lbl.setStyleSheet(
                f"background-color: {c}; color: {fg}; border-radius: 4px; "
                f"font-weight: bold; font-size: 12px; padding: 4px;"
            )
            lbl.setText(f"< {sb.value()} s")

    def _save(self):
        database.db.SetPrompterConfig(
            prompter_text_size=str(self.sb_prompter_text_size.value()),
            prompter_text_color=self.cb_prompter_text_color.get_color(),
            next_cue_text_size=str(self.sb_next_cue_text_size.value()),
            next_cue_text_color=self.cb_next_cue_text_color.get_color(),
            next_cue_outline_width=str(self.sb_next_cue_outline.value()),
            second_cue_text_size=str(self.sb_second_cue_text_size.value()),
            second_cue_text_color=self.cb_second_cue_text_color.get_color(),
            second_cue_outline_width=str(self.sb_second_cue_outline.value()),
            clock_text_size=str(self.sb_clock_text_size.value()),
            clock_text_color=self.cb_clock_text_color.get_color(),
            clock_outline_width=str(self.sb_clock_outline.value()),
            blink_first_time=str(self.sb_blink_first_time.value() * 1000),
            blink_first_color=self.cb_blink_first_color.get_color(),
            blink_second_time=str(self.sb_blink_second_time.value() * 1000),
            blink_second_color=self.cb_blink_second_color.get_color(),
            blink_third_time=str(self.sb_blink_third_time.value() * 1000),
            blink_third_color=self.cb_blink_third_color.get_color(),
        )
        self.accept()


class CreateShowDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)
        self.setWindowTitle("Créer une nouvelle emmsion")

        self.ui.show_time.setDisplayFormat("mm:ss")
        self.ui.show_time.setTime(QTime(0, 0, 0))  # valeur initiale 00:00

        self.ui.create.clicked.connect(self.create_show)
        self.ui.abort.clicked.connect(self.reject)

    def create_show(self):
        name = self.ui.show_name.text()
        desc = self.ui.show_description.text()

        time = self.ui.show_time.time()
        ms = (time.minute() * 60 + time.second()) * 1000

        if not name or not desc or ms == 0:
            QMessageBox.warning(self, "Config invalide", "Tout les champs ne sont pas remplis, ou le temps de l'emmision est nul.")
            return
        
        database.db.AddShow(name, desc, ms)
        logger.info(f"Émission ajoutée : '{name}' (durée {ms}ms)")
        self.accept()

class ModifyShowDialog(QDialog):
    def __init__(self, show_id):
        super().__init__()
        source_show =database.db.GetShowById(show_id)
        self.show_id = show_id

        self.ui = Ui_Dialog()
        self.ui.setupUi(self)
        self.setWindowTitle("Modifier une emmsion")

        self.ui.show_time.setDisplayFormat("mm:ss")
        self.ui.show_time.setTime(QTime(0, 0, 0))  # valeur initiale 00:00

        try:
            self.ui.show_name.setText(source_show["nom"])
            self.ui.show_description.setText(source_show["description"])

            total_sec = source_show["duree"] // 1000
            minutes = total_sec // 60
            seconds = total_sec % 60
            self.ui.show_time.setTime(QTime(0, minutes, seconds))

        except Exception as e:
            print(f"Erreur lors du chargement de l'emmsion pour modification: {e}")
            QMessageBox.critical(self, "Erreur", "Impossible de charger les données de l'emmsion pour modification.")
            self.reject()
            return
        
        self.ui.create.clicked.connect(self.modify_show)
        self.ui.abort.clicked.connect(self.reject)

    def modify_show(self):
        name = self.ui.show_name.text()
        desc = self.ui.show_description.text()

        time = self.ui.show_time.time()
        ms = (time.minute() * 60 + time.second()) * 1000

        if not name or not desc or ms == 0:
            QMessageBox.warning(self, "Config invalide", "Tout les champs ne sont pas remplis, ou le temps de l'emmision est nul.")
            return
        
        database.db.ModifyShow(self.show_id, name, desc, ms)
        logger.info(f"Émission modifiée (id={self.show_id}) : '{name}' (durée {ms}ms)")
        self.accept()


class ModifyOSCConfigDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.ui = Ui_OSCDialog()
        self.ui.setupUi(self)
        self.setWindowTitle("Configurer la connexion OSC")

        osc_ip, osc_port = database.db.GetOSCConfig()
        if osc_ip and osc_port:
            self.ui.osc_ip.setText(osc_ip)
            self.ui.osc_port.setValue(int(osc_port))

        self.ui.abort.clicked.connect(self.reject)
        self.ui.create.clicked.connect(self.save_config)
        
    def save_config(self):
        osc_port = self.ui.osc_port.value()
        osc_ip = self.ui.osc_ip.text()
        if not osc_ip or not osc_port:
            QMessageBox.warning(self, "Config invalide", "L'IP et le port OSC doivent être saisies.")
            return

        database.db.SetOSCConfig(osc_ip, osc_port)


        #mise a jour de la connexion OSC avec les nouvelles config
        try:
            osc.disconnect_osc_client()
            osc.connect_osc_client()
            
            osc.osc_client.send("/test", "Configuration OSC MAJ") #RETIRER APRES TEST, envoie message bidon pour tester la reconnexion
        except Exception as e:
            print(f"Erreur lors de la mise à jour de la connexion OSC: {e}")
            QMessageBox.critical(self, "Erreur", "La configuration a été sauvegardée, mais la connexion OSC n'a pas pu être mise à jour.")
            self.accept()
            return

        QMessageBox.information(self, "Config sauvegardée", "La configuration OSC a été mise à jour")
        self.accept()


class AddModifyCueDialog(QDialog):
    def __init__(self, action, show_id=None, cue_id=None):
        super().__init__()
        self.ui = Ui_CueDialog()
        self.ui.setupUi(self)
        self.setWindowTitle(f"{action} un cue")

        self.action = action
        self.show_id = show_id
        self.cue_id = cue_id

        self.ui.abort.clicked.connect(self.reject)
        self.ui.validate.clicked.connect(self.save_cue)

        from PySide6.QtGui import QShortcut, QKeySequence
        QShortcut(QKeySequence(Qt.Key.Key_Return), self).activated.connect(self.save_cue)
        QShortcut(QKeySequence(Qt.Key.Key_Enter),  self).activated.connect(self.save_cue)

        if self.action == 'modifier' and cue_id is not None:
            try:
                cue = database.db.GetCueById(cue_id)
                self.ui.cue_title.setText(cue["nom"])
                self.ui.cue_desc.setText(cue["description"])

                total_sec = cue["temps"] // 1000
                minutes = total_sec // 60
                seconds = total_sec % 60
                self.ui.time.setTime(QTime(0, minutes, seconds))

                self.ui.osc_command.setText(cue["osc_url"])
                try:
                    self.ui.osc_args.setValue(int(cue["osc_args"]))
                except:
                    self.ui.osc_args.setValue(0)
                
                #mettre ici a jour le champs de la couleure du cue 

            except Exception as e:
                print(f"Erreur lors du chargement du cue pour modification: {e}")
                QMessageBox.critical(self, "Erreur", "Impossible de charger les données du cue pour modification.")
                self.reject()
                return
        
    
    def save_cue(self):
        name = self.ui.cue_title.text()
        desc = self.ui.cue_desc.text()
        time = self.ui.time.time()
        ms = (time.minute() * 60 + time.second()) * 1000
        print(f"Temps du cue en ms: {ms}")  # Debug: Affiche le temps converti en millisecondes 
        osc_command = self.ui.osc_command.text()
        osc_args = self.ui.osc_args.value()

        if not name or ms == 0 :
            QMessageBox.warning(self, "Config invalide", "Le nom et le temps doivent être saisies.")
            return

        if self.action == 'ajouter':
            try:
                database.db.AddCueToShow(database.db.GetActiveShow(), name, desc, ms, osc_command, osc_args, "#161A16")
                logger.info(f"Cue ajouté : '{name}' à {ms}ms  |  OSC: {osc_command} {osc_args}")
                self.accept()
            except Exception as e:
                logger.error(f"Erreur ajout cue '{name}' : {e}")
                QMessageBox.critical(self, "Erreur", "Impossible d'ajouter le cue à la base de données.")
                self.reject()
                return
            
        elif self.action == 'modifier':
            try:
                database.db.ModifyCue(self.cue_id, name, desc, ms, osc_command, osc_args, "#161A16")
                logger.info(f"Cue modifié (id={self.cue_id}) : '{name}' à {ms}ms  |  OSC: {osc_command} {osc_args}")
                self.accept()
            except Exception as e:
                logger.error(f"Erreur modification cue id={self.cue_id} : {e}")
                QMessageBox.critical(self, "Erreur", "Impossible de modifier le cue dans la base de données.")
                self.reject()
                return


#renvoie l'enplacement du fichier mot de passe de l'interface 
#s'adapte si on est compilé ou non avec frozen 
#Sans ca j'avais un bug de windows qui m'enpechait d'ecrire dans C:/program files
def _get_stpass_path():
    if getattr(sys, 'frozen', False):
        base = os.path.join(os.environ.get("LOCALAPPDATA", os.path.expanduser("~")),
                            "Foxx Production", "ShowTimer")
        os.makedirs(base, exist_ok=True)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "showtimer.stpass")

#creer le fichier de mot de passe au premier demarage ou si inexistant
#pour l'instant par defaut c'est admin
#Voir pour ajouter une interface de changement du MDP ???????
def init_password_file():
    path = _get_stpass_path()
    if not os.path.exists(path):
        with open(path, "w", encoding="utf-8") as f:
            f.write("admin") #MDP DEFAUT = admin
        logger.info(f"Fichier de mot de passe créé : {path}")

#Lit le mot de passe dans le fichier 
#renvoie le mot de passe
#Il se base sur le chemin renvoyé par _get_stpass_path
def _read_password():
    with open(_get_stpass_path(), "r", encoding="utf-8") as f:
        return f.read().strip()


#Interface de verouillage 

#########GENEREE PAR IA############ !!!!!!!!!!!!

class LockDialog(QDialog):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Interface verrouillée")
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setModal(True)
        self.setFixedSize(420, 500)
        self.setStyleSheet("background-color: #1a1a1a; color: #e0e0e0;")

        layout = QVBoxLayout(self)
        layout.setSpacing(14)
        layout.setContentsMargins(32, 32, 32, 32)

        # Logo
        logo_label = QLabel()
        pixmap = QPixmap("icon/icon.png")
        if not pixmap.isNull():
            logo_label.setPixmap(pixmap.scaled(180, 180, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        logo_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(logo_label)

        # Titre
        lbl = QLabel("Interface lockée")
        lbl.setAlignment(Qt.AlignCenter)
        lbl.setStyleSheet("font-size: 18px; font-weight: bold; color: #ffffff;")
        layout.addWidget(lbl)

        lbl2 = QLabel("Saisissez le mot de passe pour unlock")
        lbl2.setAlignment(Qt.AlignCenter)
        lbl2.setStyleSheet("font-size: 11px; color: #888888;")
        layout.addWidget(lbl2)

        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.setPlaceholderText("Mot de passe...")
        self.password_input.setStyleSheet("""
            QLineEdit {
                background-color: #2a2a2a;
                border: 1px solid #3a3a3a;
                border-radius: 5px;
                padding: 8px;
                font-size: 13px;
                color: #e0e0e0;
            }
            QLineEdit:focus { border: 1px solid #0078d4; }
        """)
        self.password_input.returnPressed.connect(self._check)
        layout.addWidget(self.password_input)

        self.lbl_error = QLabel("")
        self.lbl_error.setStyleSheet("color: #ff4444; font-size: 11px;")
        self.lbl_error.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.lbl_error)

        btn = QPushButton("Déverrouiller")
        btn.setStyleSheet("""
            QPushButton {
                background-color: #0078d4;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #005a9e; }
            QPushButton:pressed { background-color: #004080; }
        """)
        btn.clicked.connect(self._check)
        layout.addWidget(btn)

        layout.addStretch()
        self.password_input.setFocus()

    def _check(self):
        if self.password_input.text() == _read_password():
            self.accept()
        else:
            self.lbl_error.setText("Mot de passe incorrect.")
            self.password_input.clear()
            self.password_input.setFocus()




####################FONCTION D'IMPORT DE DONNEES DEPUIS EXCEL/CSV####################
#Toutes ces fonction d'import de donnée et d'export sont génées par IA !!!!!!!!!!
#A reverifer en detail

def _temps_to_ms(val):
    if val is None:
        return 0

    import datetime

    # Excel stocke les heures comme float (ex: 05:30 interprété HH:MM → 0.22916...)
    # On traite HH comme MM et MM comme SS pour obtenir MM:SS
    if isinstance(val, float):
        total_s = round(val * 86400)
        m = total_s // 3600        # ce qu'Excel croit être des heures → minutes
        s = (total_s % 3600) // 60 # ce qu'Excel croit être des minutes → secondes
        return (m * 60 + s) * 1000

    # Valeur datetime.time (openpyxl peut retourner ce type)
    if isinstance(val, datetime.time):
        # val.hour = minutes, val.minute = secondes
        return (val.hour * 60 + val.minute) * 1000

    val = str(val).strip()
    parts = val.split(":")
    try:
        if len(parts) == 3:
            return (int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])) * 1000
        elif len(parts) == 2:
            return (int(parts[0]) * 60 + int(parts[1])) * 1000
        return 0
    except Exception:
        return 0


def _rows_to_cues(headers, rows):
    COLS = {"nom": "TITRE", "description": "DESCRIPTION", "temps": "TEMPS",
            "osc_url": "URL_OSC", "osc_args": "ARG_OSC"}

    def get(row, col_name):
        try:
            idx = headers.index(col_name)
            v = row[idx]
            return v if v is not None else ""
        except (ValueError, IndexError):
            return ""

    cues = []
    for row in rows:
        nom = str(get(row, COLS["nom"])).strip()
        if not nom:
            continue
        cues.append({
            "nom":         nom,
            "description": str(get(row, COLS["description"])).strip(),
            "temps":       _temps_to_ms(get(row, COLS["temps"])),
            "osc_url":     str(get(row, COLS["osc_url"])).strip(),
            "osc_args":    str(get(row, COLS["osc_args"])).strip(),
        })
    return cues


def parse_file_to_cues(filepath):
    ext = filepath.rsplit(".", 1)[-1].lower()

    if ext in ("xlsx", "xls"):
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]

    elif ext == "csv":
        import csv
        with open(filepath, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f, delimiter=";")
            all_rows = list(reader)
        if not all_rows:
            return []
        headers = [h.strip() for h in all_rows[0]]
        rows = all_rows[1:]

    else:
        raise ValueError(f"Format non supporté : {ext}")

    return _rows_to_cues(headers, rows)


class ImportExcelDialog(QDialog):
    def __init__(self, parent=None, mode="excel"):
        super().__init__(parent)
        label = "CSV" if mode == "csv" else "Excel"
        self.setWindowTitle(f"Importer des cues depuis {label}")
        self.setMinimumSize(900, 500)
        self.cues = []

        layout = QVBoxLayout(self)

        # Sélection du fichier
        top = QHBoxLayout()
        self.path_input = QLineEdit()
        self.path_input.setPlaceholderText(f"Chemin vers le fichier {label}...")
        self.path_input.setReadOnly(True)
        top.addWidget(self.path_input)
        btn_browse = QPushButton("Parcourir...")
        btn_browse.clicked.connect(self._browse)
        top.addWidget(btn_browse)
        layout.addLayout(top)

        # Tableau de preview des cues qu'on va importer
        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["Nom", "Description", "Temps", "URL OSC", "Args OSC"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        layout.addWidget(self.table)

        self.lbl_count = QLabel("Aucun fichier chargé.")
        layout.addWidget(self.lbl_count)

        # Boutons
        btn_box = QDialogButtonBox()
        self.btn_import = btn_box.addButton("Importer", QDialogButtonBox.AcceptRole)
        self.btn_import.setEnabled(False)
        btn_box.addButton("Annuler", QDialogButtonBox.RejectRole)
        btn_box.accepted.connect(self._import)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def _browse(self):
        filtre = getattr(self, "_browse_filter", "Tous les formats supportés (*.xlsx *.xls *.csv);;Excel (*.xlsx *.xls);;CSV (*.csv)")
        path, _ = QFileDialog.getOpenFileName(self, "Ouvrir un fichier", "", filtre)
        if not path:
            return
        self.path_input.setText(path)
        try:
            self.cues = parse_file_to_cues(path)
            self._fill_table()
            self.lbl_count.setText(f"{len(self.cues)} cue(s) trouvé(s).")
            self.btn_import.setEnabled(len(self.cues) > 0)
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Impossible de lire le fichier :\n{e}")

    def _fill_table(self):
        self.table.setRowCount(0)
        RED = QColor("#FF4444")
        for cue in self.cues:
            row = self.table.rowCount()
            self.table.insertRow(row)
            ms = cue["temps"]
            h = ms // 3600000
            m = (ms % 3600000) // 60000
            s = (ms % 60000) // 1000
            values = [
                cue["nom"],
                cue["description"],
                f"{h:02d}:{m:02d}:{s:02d}",
                cue["osc_url"],
                cue["osc_args"],
            ]
            for col, val in enumerate(values):
                item = QTableWidgetItem(val)
                if not val.strip() or (col == 2 and ms == 0):
                    item.setBackground(RED)
                self.table.setItem(row, col, item)

    def _import(self):
        """Relit le tableau (modifiable) et met à jour self.cues avant d'accepter."""
        self.cues = []
        for row in range(self.table.rowCount()):
            def cell(col):
                item = self.table.item(row, col)
                return item.text().strip() if item else ""

            temps_str = cell(2)
            parts = temps_str.split(":")
            try:
                if len(parts) == 3:
                    ms = (int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])) * 1000
                elif len(parts) == 2:
                    ms = (int(parts[0]) * 60 + int(parts[1])) * 1000
                else:
                    ms = 0
            except Exception:
                ms = 0

            nom = cell(0)
            if not nom:
                continue
            self.cues.append({
                "nom":         nom,
                "description": cell(1),
                "temps":       ms,
                "osc_url":     cell(3),
                "osc_args":    cell(4),
            })
        self.accept()


def _ms_to_hms(ms):
    ms = max(0, int(ms))
    h  = ms // 3600000
    m  = (ms % 3600000) // 60000
    s  = (ms % 60000) // 1000
    return f"{h:02d}:{m:02d}:{s:02d}"


def export_show_to_pdf(show_id, output_path):
    """Exporte le show (show_id) et tous ses cues en PDF vers output_path."""
    import base64, os
    from PySide6.QtGui import QTextDocument, QPageSize, QPageLayout
    from PySide6.QtPrintSupport import QPrinter
    from PySide6.QtCore import QSizeF

    show = database.db.GetShowById(show_id)
    if show is None:
        raise ValueError(f"Show {show_id} introuvable.")

    cues = database.db.GetAllCuesFromShow(show_id) or []
    cues = sorted(cues, key=lambda c: c["temps"])

    # ── Icône en base64 ──
    icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "icon", "icon.png")
    icon_src = ""
    if os.path.exists(icon_path):
        with open(icon_path, "rb") as f:
            icon_src = "data:image/png;base64," + base64.b64encode(f.read()).decode()

    logo_html = f'<img src="{icon_src}" height="48" style="vertical-align:middle;">' if icon_src else ""
    logo_small = logo_html.replace('height="48"', 'height="20"') if logo_html else ""

    # ── Infos show ──
    total_ms   = show.get("duree", 0)
    total_str  = _ms_to_hms(total_ms)
    show_title = show.get("nom", "")
    show_desc  = show.get("description", "") or "—"

    # ── Lignes du tableau de cues ──
    rows_html = ""
    prev_ms = 0
    for i, cue in enumerate(cues):
        t_ms   = cue.get("temps", 0)
        delta  = t_ms - prev_ms if i > 0 else t_ms
        prev_ms = t_ms

        bg = "#ffffff" if i % 2 == 0 else "#f5f5f5"
        rows_html += f"""
        <tr style="background:{bg};">
            <td style="text-align:center;">{i + 1}</td>
            <td><b>{cue.get('nom','')}</b></td>
            <td class="desc">{cue.get('description','') or '&mdash;'}</td>
            <td style="text-align:center;">{_ms_to_hms(t_ms)}</td>
            <td style="text-align:center;">+{_ms_to_hms(delta)}</td>
            <td style="font-family:monospace;">{cue.get('osc_url','') or '&mdash;'}</td>
            <td style="font-family:monospace;">{cue.get('osc_args','') or '&mdash;'}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  body      {{ font-family: Arial, sans-serif; font-size: 10pt; color: #1a1a1a; margin: 0; padding: 0; }}
  .page-header {{
      border-bottom: 3px solid #2c3e50; padding-bottom: 8px; margin-bottom: 18px;
  }}
  .page-header table {{ width: auto; border-collapse: collapse; }}
  .page-header td {{ padding: 0 10px 0 0; vertical-align: middle; border: none; background: none; }}
  .page-header h1 {{ margin: 0; font-size: 18pt; color: #2c3e50; }}
  .show-block {{
      background: #f0f4f8; border-left: 5px solid #2c3e50;
      padding: 10px 14px; margin-bottom: 20px; border-radius: 3px;
  }}
  .show-block h2 {{ margin: 0 0 4px 0; font-size: 14pt; color: #2c3e50; }}
  .show-block p  {{ margin: 2px 0; font-size: 9pt; color: #444; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 6.5pt; table-layout: fixed; }}
  thead tr {{ background: #2c3e50; color: white; }}
  thead th {{ padding: 3px 4px; text-align: left; font-weight: bold; overflow: hidden; }}
  tbody td {{ padding: 2px 4px; border-bottom: 1px solid #ddd; vertical-align: top; overflow: hidden; word-wrap: break-word; }}
  tbody td.desc {{ white-space: normal; }}
  .footer {{
      margin-top: 24px; border-top: 1px solid #ccc; padding-top: 6px;
      font-size: 8pt; color: #888;
  }}
</style>
</head>
<body>

<div class="page-header">
  <table><tr>
    <td>{logo_html}</td>
    <td><h1>Rapport &mdash; ShowTimer V1.0</h1></td>
  </tr></table>
</div>

<br>

<div class="show-block">
  <h2>{show_title}</h2>
  <p><b>Description :</b> {show_desc}</p>
  <p><b>Dur&eacute;e totale :</b> {total_str}</p>
  <p><b>Nombre de cues :</b> {len(cues)}</p>
</div>

<br>


<table>
  <thead>
    <tr>
      <th width="3%">N&deg;</th>
      <th width="13%">Nom</th>
      <th width="24%" class="desc">Description</th>
      <th width="10%">D&eacute;clenchement</th>
      <th width="10%">&Delta; Pr&eacute;c&eacute;dent</th>
      <th width="27%">Commande OSC</th>
      <th width="13%">Arg. OSC</th>
    </tr>
  </thead>
  <tbody>
    {rows_html}
  </tbody>
</table>

<div class="footer">
  {logo_small}
  &copy; Foxx Prod 2026 &mdash; ShowTimer V1.0
</div>

</body>
</html>"""

    printer = QPrinter(QPrinter.HighResolution)
    printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
    printer.setOutputFileName(output_path)
    printer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
    printer.setPageOrientation(QPageLayout.Orientation.Portrait)

    doc = QTextDocument()
    doc.setHtml(html)
    doc.setPageSize(QSizeF(printer.pageRect(QPrinter.Unit.Point).size()))
    doc.print_(printer)
#########FIN GENERE PAR IA############ !!!!!!!!!!!!
